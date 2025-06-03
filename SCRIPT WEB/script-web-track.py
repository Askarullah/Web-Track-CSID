# app.py - Main Flask application
from flask import Flask, render_template, request, jsonify, session, send_file
import pandas as pd
import os
import uuid
from werkzeug.utils import secure_filename
import tempfile
import json
from datetime import datetime
import io
import openpyxl

# Use os.path.join for better cross-platform compatibility
app = Flask(__name__, template_folder=os.path.join('..', 'templates'))
app.secret_key = 'your-secret-key-change-this'  # Change this in production
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Global storage for workbook data (in production, use Redis or database)
workbook_storage = {}

def find_in_all_sheets(session_id, search_value):
    """Find search value in all sheets for a specific session"""
    if session_id not in workbook_storage:
        return "No data loaded"
    
    workbook_data = workbook_storage[session_id]['data']
    found_sheets = []
    search_str = str(search_value)
    
    for sheet_name, column_data in workbook_data.items():
        str_data = [str(val) for val in column_data if str(val) != 'nan']
        if search_str in str_data:
            found_sheets.append(sheet_name)
    
    if found_sheets:
        return ", ".join(found_sheets)
    else:
        return "Not Found"

@app.route('/')
def index():
    """Main page"""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    return render_template('home.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle Excel file upload"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Read Excel file
        workbook_data = {}
        with pd.ExcelFile(file) as xls:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if len(df.columns) >= 4:  # Check if column D exists
                    column_d = df.iloc[:, 3].astype(str).dropna()  # Column D (index 3)
                    workbook_data[sheet_name] = column_d.tolist()
                else:
                    workbook_data[sheet_name] = []
        
        # Store in session storage
        session_id = session['session_id']
        workbook_storage[session_id] = {
            'data': workbook_data,
            'filename': secure_filename(file.filename),
            'upload_time': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'message': f'Loaded data from {len(workbook_data)} sheets',
            'sheets': list(workbook_data.keys())
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to load Excel file: {str(e)}'}), 500

@app.route('/add_csid', methods=['POST'])
def add_csid():
    """Add single CSID"""
    data = request.get_json()
    csid = data.get('csid', '').strip()
    
    if not csid:
        return jsonify({'error': 'CSID cannot be empty'}), 400
    
    session_id = session['session_id']
    found_sheets = find_in_all_sheets(session_id, csid)
    
    return jsonify({
        'success': True,
        'csid': csid,
        'found_sheets': found_sheets
    })

#SCRIPT YANG TRIGGER TRACKING CO
@app.route('/add_bulk_csids', methods=['POST'])
def add_bulk_csids():
    """Add multiple CSIDs"""
    data = request.get_json()
    csids_text = data.get('csids', '').strip()
    
    if not csids_text:
        return jsonify({'error': 'No CSIDs provided'}), 400
    
    # Parse CSIDs
    import re
    csids = re.split(r'[\s,;\n\t]+', csids_text)
    csids = [csid.strip() for csid in csids if csid.strip()]
    
    if not csids:
        return jsonify({'error': 'No valid CSIDs found'}), 400
    
    session_id = session['session_id']
    results = []
    
    for csid in csids:
        found_sheets = find_in_all_sheets(session_id, csid)
        results.append({
            'csid': csid,
            'found_sheets': found_sheets
        })
    
    return jsonify({
        'success': True,
        'results': results,
        'count': len(results)
    })

@app.route('/refresh_csids', methods=['POST'])
def refresh_csids():
    """Refresh existing CSIDs"""
    data = request.get_json()
    csids = data.get('csids', [])
    
    if not csids:
        return jsonify({'error': 'No CSIDs to refresh'}), 400
    
    session_id = session['session_id']
    results = []
    
    for csid in csids:
        found_sheets = find_in_all_sheets(session_id, csid)
        results.append({
            'csid': csid,
            'found_sheets': found_sheets
        })
    
    return jsonify({
        'success': True,
        'results': results
    })

@app.route('/export', methods=['POST'])
def export_data():
    """Export data to Excel"""
    data = request.get_json()
    export_data = data.get('data', [])
    
    if not export_data:
        return jsonify({'error': 'No data to export'}), 400
    
    try:
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='CSID_Results')
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'csid_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

@app.route('/status')
def get_status():
    """Get current session status"""
    session_id = session.get('session_id')
    
    if session_id in workbook_storage:
        wb_info = workbook_storage[session_id]
        return jsonify({
            'loaded': True,
            'filename': wb_info['filename'],
            'sheets': list(wb_info['data'].keys()),
            'upload_time': wb_info['upload_time']
        })
    else:
        return jsonify({'loaded': False})

@app.route('/tracking-co.html')
def track_home():
    return render_template('tracking-co.html')

@app.route('/tracking-odp.html')
def track_odp():
    return render_template('tracking-odp.html')
    
@app.route('/tracking-ip.html')
def track_ip () :
    return render_template('tracking-ip.html')

#script buat tracking-odp.html
@app.route('/search-odp', methods=['POST'])
def search_odp():
    odp_id = request.form.get('odp_id')
    file = request.files['file']
    
    wb = openpyxl.load_workbook(file)
    results = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "TRACK ODP":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and str(row[1]).strip() == odp_id.strip():
                results.append({
                    'ip': row[2],
                    'csid': row[3]
                })
    
    return jsonify(results)

#script buat tracking-ip.html
@app.route('/search-ip', methods=['POST'])
def search_ip():
    csid_input = request.form.get('csid')
    file = request.files['file']
    
    # Parse multiple CSIDs - split by newlines, commas, or semicolons
    csids = []
    if csid_input:
        # Split by various delimiters and clean up
        import re
        csid_list = re.split(r'[,;\n\r]+', csid_input.strip())
        csids = [csid.strip() for csid in csid_list if csid.strip()]
    
    if not csids:
        return jsonify({'error': 'Please provide at least one CSID'})
    
    wb = openpyxl.load_workbook(file)
    results = []
    found_csids = set()
    
    # Search through all sheets except "TRACK ODP"
    for sheet_name in wb.sheetnames:
        if sheet_name == "TRACK ODP":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3]:  # Check if CSID column has value
                row_csid = str(row[3]).strip()
                # Check if this row's CSID matches any of the searched CSIDs
                for search_csid in csids:
                    if row_csid == search_csid:
                        results.append({
                            'csid': row_csid,
                            'ip': row[2] if row[2] else 'N/A',
                            'sheet': sheet_name
                        })
                        found_csids.add(search_csid)
                        break
    
    # Add entries for CSIDs that were not found
    not_found_csids = set(csids) - found_csids
    for csid in not_found_csids:
        results.append({
            'csid': csid,
            'ip': 'Not Found',
            'sheet': 'N/A'
        })
    
    # Sort results to show found ones first, then not found
    results.sort(key=lambda x: (x['ip'] == 'Not Found', x['csid']))
    
    return jsonify(results)


    
@app.route('/reset_file', methods=['POST'])
def reset_file():
    """Reset the uploaded file data"""
    session_id = session.get('session_id')
    if session_id in workbook_storage:
        del workbook_storage[session_id]
    return jsonify({'success': True, 'message': 'File data reset successfully'})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5050)